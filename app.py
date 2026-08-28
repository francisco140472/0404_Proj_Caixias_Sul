import os
import re
import io
import ssl
import math
import json
import hashlib
import traceback
import smtplib
import shutil
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import pandas as pd
import pyodbc
import pdfplumber
import unicodedata
from collections import Counter
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from concurrent.futures import ProcessPoolExecutor, as_completed

# ==========================================
# CONFIGURAÇÃO DE REDE / PROXY CORPORATIVO
# ==========================================
try:
    ssl._create_default_https_context = ssl._create_unverified_context
except AttributeError:
    pass

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

try:
    import contextily as ctx
    import pyproj
    HAS_CONTEXTILY = True
except ImportError:
    HAS_CONTEXTILY = False

from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth

app = Flask(__name__)
app.secret_key = 'chave_secreta_super_segura_para_sessoes'  # Necessário para gerenciar o login

UPLOAD_FOLDER = 'uploads'
DATA_FOLDER = 'data'
PDFS_CAMPO_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pdfs_campo')
FUNCIONARIOS_JSON = os.path.join(DATA_FOLDER, 'funcionarios.json')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DATA_FOLDER, exist_ok=True)
os.makedirs(PDFS_CAMPO_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Configuração opcional de e-mail (variáveis de ambiente ou preenchimento manual)
EMAIL_CONFIG = {
    'smtp_server': os.environ.get('SMTP_SERVER', ''),
    'smtp_port': int(os.environ.get('SMTP_PORT', '587')),
    'smtp_user': os.environ.get('SMTP_USER', ''),
    'smtp_password': os.environ.get('SMTP_PASSWORD', ''),
    'smtp_from': os.environ.get('SMTP_FROM', ''),
    'use_tls': os.environ.get('SMTP_USE_TLS', 'true').lower() == 'true',
}
SMTP_CONFIG_FILE = os.path.join(DATA_FOLDER, 'smtp_config.json')


def carregar_config_email_arquivo():
    """Carrega configuração SMTP opcional de data/smtp_config.json."""
    if not os.path.exists(SMTP_CONFIG_FILE):
        return
    try:
        with open(SMTP_CONFIG_FILE, 'r', encoding='utf-8') as arquivo:
            cfg = json.load(arquivo)
        for chave in ('smtp_server', 'smtp_user', 'smtp_password', 'smtp_from'):
            if cfg.get(chave):
                EMAIL_CONFIG[chave] = str(cfg[chave]).strip()
        if cfg.get('smtp_port'):
            EMAIL_CONFIG['smtp_port'] = int(cfg['smtp_port'])
        if 'use_tls' in cfg:
            EMAIL_CONFIG['use_tls'] = bool(cfg['use_tls'])
    except Exception as exc:
        print(f'[AVISO] Erro ao ler {SMTP_CONFIG_FILE}: {exc}')


carregar_config_email_arquivo()

_FUNCIONARIOS_USA_SQL = False
_APP_INICIALIZADO = False

# ==========================================
# CONFIGURAÇÃO DOS BANCOS DE DADOS SQL SERVER
# ==========================================
DB_CONFIG = {
    'driver': '{ODBC Driver 17 for SQL Server}', 
    'server': '200.98.80.97', 
    'uid': 'sa',                     
    'pwd': 'SantoAndre2021'                  
}

def obter_conexao_sql(nome_banco):
    """Conecta dinamicamente ao banco especificado: E_GLOBAL_COMERCIAL ou SABESP_LIGACOES"""
    string_conexao = (
        f"DRIVER={DB_CONFIG['driver']};"
        f"SERVER={DB_CONFIG['server']};"
        f"DATABASE={nome_banco};"
        f"UID={DB_CONFIG['uid']};"
        f"PWD={DB_CONFIG['pwd']};"
    )
    return pyodbc.connect(string_conexao)


def inicializar_tabela_funcionarios():
    """Cria a tabela de funcionários no SQL Server, se ainda não existir."""
    global _FUNCIONARIOS_USA_SQL
    try:
        conn = obter_conexao_sql('SABESP_LIGACOES')
        cursor = conn.cursor()
        cursor.execute("""
            IF NOT EXISTS (
                SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'FUNCIONARIOS'
            )
            BEGIN
                CREATE TABLE dbo.FUNCIONARIOS (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    nome NVARCHAR(200) NOT NULL,
                    matricula NVARCHAR(50) NULL,
                    cargo NVARCHAR(100) NULL,
                    email NVARCHAR(200) NULL,
                    telefone NVARCHAR(30) NULL,
                    setor NVARCHAR(100) NULL,
                    ativo CHAR(1) NOT NULL DEFAULT 'S',
                    data_cadastro DATETIME NOT NULL DEFAULT GETDATE(),
                    data_atualizacao DATETIME NULL
                )
            END
        """)
        conn.commit()
        conn.close()
        _FUNCIONARIOS_USA_SQL = True
        print('[OK] Módulo de funcionários usando SQL Server (SABESP_LIGACOES.dbo.FUNCIONARIOS)')
    except Exception as exc:
        _FUNCIONARIOS_USA_SQL = False
        print(f'[AVISO] Funcionários em arquivo local (JSON): {exc}')


def _ler_funcionarios_json():
    """Lê funcionários do arquivo JSON local."""
    if not os.path.exists(FUNCIONARIOS_JSON):
        return []
    with open(FUNCIONARIOS_JSON, 'r', encoding='utf-8') as arquivo:
        return json.load(arquivo)


def _gravar_funcionarios_json(lista):
    """Grava funcionários no arquivo JSON local."""
    with open(FUNCIONARIOS_JSON, 'w', encoding='utf-8') as arquivo:
        json.dump(lista, arquivo, ensure_ascii=False, indent=2)


def _funcionario_para_dict(row):
    """Normaliza registro de funcionário para dicionário."""
    if isinstance(row, dict):
        return row
    return {
        'id': row[0],
        'nome': row[1],
        'matricula': row[2] or '',
        'cargo': row[3] or '',
        'email': row[4] or '',
        'telefone': row[5] or '',
        'setor': row[6] or '',
        'ativo': row[7],
        'data_cadastro': row[8].isoformat() if row[8] else None,
        'data_atualizacao': row[9].isoformat() if row[9] else None,
    }


def listar_funcionarios(apenas_ativos=True):
    """Lista funcionários cadastrados."""
    if _FUNCIONARIOS_USA_SQL:
        conn = obter_conexao_sql('SABESP_LIGACOES')
        cursor = conn.cursor()
        sql = """
            SELECT id, nome, matricula, cargo, email, telefone, setor, ativo,
                   data_cadastro, data_atualizacao
            FROM dbo.FUNCIONARIOS
        """
        if apenas_ativos:
            sql += " WHERE ativo = 'S'"
        sql += ' ORDER BY nome'
        cursor.execute(sql)
        rows = cursor.fetchall()
        conn.close()
        return [_funcionario_para_dict(r) for r in rows]

    lista = _ler_funcionarios_json()
    if apenas_ativos:
        lista = [f for f in lista if f.get('ativo', 'S') == 'S']
    return sorted(lista, key=lambda f: f.get('nome', ''))


def _para_int(valor, mensagem='Identificador inválido.'):
    """Converte valor para inteiro com mensagem amigável."""
    if valor is None or valor == '':
        raise ValueError(mensagem)
    try:
        return int(valor)
    except (TypeError, ValueError) as exc:
        raise ValueError(mensagem) from exc


def _extrair_id_inserido_sql(cursor):
    """Obtém o ID gerado após INSERT no SQL Server."""
    row = cursor.fetchone()
    if row and row[0] is not None:
        return _para_int(row[0], 'Não foi possível confirmar o cadastro do funcionário.')

    cursor.execute('SELECT CAST(SCOPE_IDENTITY() AS INT)')
    row = cursor.fetchone()
    if row and row[0] is not None:
        return _para_int(row[0], 'Não foi possível confirmar o cadastro do funcionário.')

    raise ValueError('Não foi possível confirmar o cadastro do funcionário. Tente novamente.')


def obter_funcionario_por_id(funcionario_id):
    """Busca um funcionário pelo identificador."""
    funcionario_id = _para_int(funcionario_id, 'Funcionário não encontrado.')
    if _FUNCIONARIOS_USA_SQL:
        conn = obter_conexao_sql('SABESP_LIGACOES')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, nome, matricula, cargo, email, telefone, setor, ativo,
                   data_cadastro, data_atualizacao
            FROM dbo.FUNCIONARIOS WHERE id = ?
        """, (funcionario_id,))
        row = cursor.fetchone()
        conn.close()
        return _funcionario_para_dict(row) if row else None

    for func in _ler_funcionarios_json():
        if int(func.get('id', 0)) == funcionario_id:
            return func
    return None


def salvar_funcionario(dados, funcionario_id=None):
    """Insere ou atualiza um funcionário."""
    nome = dados.get('nome', '').strip()
    if not nome:
        raise ValueError('Nome do funcionário é obrigatório.')

    matricula = dados.get('matricula', '').strip()
    cargo = dados.get('cargo', '').strip()
    email = dados.get('email', '').strip()
    telefone = dados.get('telefone', '').strip()
    setor = dados.get('setor', '').strip()
    ativo = 'S' if dados.get('ativo', True) in (True, 'S', 's', '1', 1) else 'N'

    if _FUNCIONARIOS_USA_SQL:
        conn = obter_conexao_sql('SABESP_LIGACOES')
        cursor = conn.cursor()
        if funcionario_id:
            cursor.execute("""
                UPDATE dbo.FUNCIONARIOS
                SET nome = ?, matricula = ?, cargo = ?, email = ?, telefone = ?,
                    setor = ?, ativo = ?, data_atualizacao = GETDATE()
                WHERE id = ?
            """, (nome, matricula, cargo, email, telefone, setor, ativo, int(funcionario_id)))
        else:
            cursor.execute("""
                INSERT INTO dbo.FUNCIONARIOS
                (nome, matricula, cargo, email, telefone, setor, ativo)
                OUTPUT INSERTED.id
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (nome, matricula, cargo, email, telefone, setor, ativo))
            novo_id = _extrair_id_inserido_sql(cursor)
        conn.commit()
        conn.close()
        return obter_funcionario_por_id(funcionario_id or novo_id)

    lista = _ler_funcionarios_json()
    agora = datetime.now().isoformat()
    if funcionario_id:
        funcionario_id = int(funcionario_id)
        encontrado = False
        for func in lista:
            if int(func.get('id', 0)) == funcionario_id:
                func.update({
                    'nome': nome, 'matricula': matricula, 'cargo': cargo,
                    'email': email, 'telefone': telefone, 'setor': setor,
                    'ativo': ativo, 'data_atualizacao': agora
                })
                encontrado = True
                break
        if not encontrado:
            raise ValueError('Funcionário não encontrado.')
    else:
        novo_id = max([int(f.get('id', 0)) for f in lista], default=0) + 1
        novo = {
            'id': novo_id, 'nome': nome, 'matricula': matricula, 'cargo': cargo,
            'email': email, 'telefone': telefone, 'setor': setor, 'ativo': ativo,
            'data_cadastro': agora, 'data_atualizacao': None
        }
        lista.append(novo)
    _gravar_funcionarios_json(lista)
    return obter_funcionario_por_id(funcionario_id or novo_id)


def remover_funcionario(funcionario_id):
    """Remove funcionário (exclusão lógica no SQL ou remoção no JSON)."""
    funcionario_id = int(funcionario_id)
    if _FUNCIONARIOS_USA_SQL:
        conn = obter_conexao_sql('SABESP_LIGACOES')
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE dbo.FUNCIONARIOS
            SET ativo = 'N', data_atualizacao = GETDATE()
            WHERE id = ?
        """, (funcionario_id,))
        conn.commit()
        conn.close()
        return True

    lista = _ler_funcionarios_json()
    nova_lista = [f for f in lista if int(f.get('id', 0)) != funcionario_id]
    if len(nova_lista) == len(lista):
        return False
    _gravar_funcionarios_json(nova_lista)
    return True


def _montar_linhas_excel(features, funcionario=None):
    """Monta linhas da planilha com coluna de responsável."""
    linhas = []
    nome_resp = funcionario.get('nome', '') if funcionario else ''
    for feat in features:
        props = feat['properties']
        inscricao = props.get('Inscricao_ZSQL')
        consumidor = props.get('Nome_Consumidor')
        for serv in props.get('Servicos', []):
            linhas.append({
                'Responsavel': nome_resp,
                'Inscricao_ZSQL': inscricao,
                'Nome_Consumidor': consumidor,
                'OS_Numero': serv.get('OS_Numero'),
                'Tipo_TSS': serv.get('Tipo'),
                'Hidrometro': serv.get('Hidrometro'),
                'Arquivo': serv.get('Nome_do_Arquivo')
            })
    return linhas


def gerar_excel_personalizado(features, funcionario=None, usuario_logado=''):
    """Gera Excel com aba de resumo e aba de ligações."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    linhas = _montar_linhas_excel(features, funcionario)
    if not linhas:
        raise ValueError('Nenhum registro para exportar.')

    df = pd.DataFrame(linhas)
    output = io.BytesIO()

    titulo_fill = PatternFill('solid', fgColor='C82333')
    header_fill = PatternFill('solid', fgColor='1E293B')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    titulo_font = Font(color='FFFFFF', bold=True, size=14)
    label_font = Font(bold=True, color='334155')
    border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Ligacoes', startrow=1)
        ws_dados = writer.sheets['Ligacoes']

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws_dados.cell(row=2, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for row in ws_dados.iter_rows(min_row=3, max_row=ws_dados.max_row, min_col=1, max_col=ws_dados.max_column):
            for cell in row:
                cell.border = border

        for col in ws_dados.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws_dados.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

        ws_resumo = writer.book.create_sheet('Resumo', 0)
        ws_resumo.merge_cells('A1:D1')
        c1 = ws_resumo['A1']
        c1.value = 'ENORSUL — Relatório Operacional de Ligações'
        c1.fill = titulo_fill
        c1.font = titulo_font
        c1.alignment = Alignment(horizontal='center', vertical='center')
        ws_resumo.row_dimensions[1].height = 28

        agora = datetime.now().strftime('%d/%m/%Y %H:%M')
        info = [
            ('Data de geração', agora),
            ('Gerado por', usuario_logado or '—'),
            ('Total de registros', len(linhas)),
            ('Total de ligações', len(features)),
        ]
        if funcionario:
            info.extend([
                ('Funcionário destinatário', funcionario.get('nome', '—')),
                ('Matrícula', funcionario.get('matricula', '—') or '—'),
                ('Cargo', funcionario.get('cargo', '—') or '—'),
                ('Setor', funcionario.get('setor', '—') or '—'),
                ('E-mail', funcionario.get('email', '—') or '—'),
                ('Telefone', funcionario.get('telefone', '—') or '—'),
            ])

        linha = 3
        for rotulo, valor in info:
            ws_resumo.cell(row=linha, column=1, value=rotulo).font = label_font
            ws_resumo.cell(row=linha, column=2, value=valor)
            linha += 1

        ws_resumo.column_dimensions['A'].width = 28
        ws_resumo.column_dimensions['B'].width = 50

    output.seek(0)
    return output


def enviar_email_smtp(destinatario, assunto, corpo, anexo_bytes, nome_arquivo):
    """Envia e-mail com anexo Excel via SMTP."""
    if not EMAIL_CONFIG.get('smtp_server') or not EMAIL_CONFIG.get('smtp_from'):
        raise ValueError('SMTP não configurado.')

    msg = MIMEMultipart()
    msg['From'] = EMAIL_CONFIG['smtp_from']
    msg['To'] = destinatario
    msg['Subject'] = assunto
    msg.attach(MIMEText(corpo, 'plain', 'utf-8'))

    conteudo = anexo_bytes.read()
    anexo_bytes.seek(0)
    parte_anexo = MIMEApplication(conteudo, _subtype='xlsx')
    parte_anexo.add_header('Content-Disposition', 'attachment', filename=nome_arquivo)
    msg.attach(parte_anexo)

    with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'], timeout=30) as servidor:
        if EMAIL_CONFIG.get('use_tls'):
            servidor.starttls()
        if EMAIL_CONFIG.get('smtp_user'):
            servidor.login(EMAIL_CONFIG['smtp_user'], EMAIL_CONFIG['smtp_password'])
        servidor.send_message(msg)


def enviar_email_outlook(destinatario, assunto, corpo, anexo_bytes, nome_arquivo):
    """Envia e-mail usando o Outlook instalado no Windows (sem SMTP)."""
    import tempfile
    try:
        import win32com.client
    except ImportError as exc:
        raise ValueError('Outlook indisponível (instale pywin32 ou configure SMTP).') from exc

    conteudo = anexo_bytes.read()
    anexo_bytes.seek(0)

    caminho_temp = os.path.join(tempfile.gettempdir(), nome_arquivo)
    with open(caminho_temp, 'wb') as arquivo:
        arquivo.write(conteudo)

    try:
        outlook = win32com.client.Dispatch('Outlook.Application')
        mail = outlook.CreateItem(0)
        mail.To = destinatario
        mail.Subject = assunto
        mail.Body = corpo
        mail.Attachments.Add(caminho_temp)
        mail.Send()
    finally:
        if os.path.exists(caminho_temp):
            try:
                os.remove(caminho_temp)
            except OSError:
                pass


def enviar_email_com_anexo(destinatario, assunto, corpo, anexo_bytes, nome_arquivo):
    """Envia e-mail para qualquer endereço cadastrado (SMTP ou Outlook)."""
    if not destinatario:
        raise ValueError('Funcionário sem e-mail cadastrado.')

    destinatario = str(destinatario).strip()
    erros = []

    if EMAIL_CONFIG.get('smtp_server') and EMAIL_CONFIG.get('smtp_from'):
        try:
            enviar_email_smtp(destinatario, assunto, corpo, anexo_bytes, nome_arquivo)
            return 'smtp'
        except Exception as exc:
            erros.append(f'SMTP: {exc}')
            anexo_bytes.seek(0)

    try:
        enviar_email_outlook(destinatario, assunto, corpo, anexo_bytes, nome_arquivo)
        return 'outlook'
    except Exception as exc:
        erros.append(f'Outlook: {exc}')

    raise ValueError(
        'Não foi possível enviar o e-mail automaticamente. '
        'Configure data/smtp_config.json ou use o Outlook no Windows. '
        + ' | '.join(erros)
    )


def requer_login():
    """Verifica se o usuário está autenticado."""
    return 'usuario' in session


@app.before_request
def inicializar_aplicacao():
    """Garante criação da tabela de funcionários na primeira requisição."""
    global _APP_INICIALIZADO
    if not _APP_INICIALIZADO:
        inicializar_tabela_funcionarios()
        _APP_INICIALIZADO = True

def remover_acentos(texto):
    if not texto:
        return ""
    return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')

def extrair_campo(padrao, texto):
    busca = re.search(padrao, texto, re.IGNORECASE)
    if busca:
        valor = busca.group(1).replace("|", "").strip()
        return valor
    return "Nao encontrado"

def processar_um_pdf(caminho_temp, nome_arquivo):
    lista_registros = []
    try:
        with pdfplumber.open(caminho_temp) as pdf:
            if not pdf.pages:
                return []
            texto = pdf.pages[0].extract_text()
            
            if texto:
                linha = {"Nome_do_Arquivo": nome_arquivo}
                
                busca_zsql = re.search(r"Z\s*/\s*S\s*/\s*Q\s*/\s*L\s*[:\s]*(\d+)\D+(\d+)\D+(\d+)\D+(\d+)", texto, re.IGNORECASE)
                if not busca_zsql:
                    busca_zsql = re.search(r"Z/S/Q/L:\s*(\d+)\s+(\d+)\s+(\d+)\s+(\d+)", texto)
                if not busca_zsql:
                    busca_zsql = re.search(r"(\d{2})[^\d](\d{2})[^\d](\d{4})[^\d](\d{3})", texto)

                if busca_zsql:
                    zona = busca_zsql.group(1).zfill(2)    
                    setor = busca_zsql.group(2).zfill(2)   
                    quadra = busca_zsql.group(3).zfill(4)  
                    lote = busca_zsql.group(4).zfill(3)    
                    linha["Inscricao_ZSQL"] = f"{zona}{setor}{quadra}{lote}"
                else:
                    linha["Inscricao_ZSQL"] = "Nao encontrado"
                    
                linha["OS_Numero"] = extrair_campo(r"ORDEM DE SERVI[CÇ]O N[UÚ]MERO:\s*\|?\s*(\d+)", texto)
                linha["Tipo"] = extrair_campo(r"TIPO:\s*\|?\s*([^\n]+)", texto)
                linha["Hidrometro"] = extrair_campo(r"HIDR[OÔ]METRO:\s*\|?\s*([A-Z0-9]+)", texto)

                consumidor_bruto = extrair_campo(r"CONSUMIDOR:\s*\|?\s*([^\n]+)", texto)
                if consumidor_bruto != "Nao encontrado":
                    consumidor_limpo = re.sub(r"\s+SITUA[CÇ][AÃ]O.*", "", consumidor_bruto, flags=re.IGNORECASE)
                    busca_separacao = re.search(r"(\d+-\d+)\s+(.*)", consumidor_limpo)
                    if busca_separacao:
                        linha["Nome_Consumidor"] = busca_separacao.group(2).strip() 
                    else:
                        linha["Nome_Consumidor"] = consumidor_limpo.strip()
                else:
                    linha["Nome_Consumidor"] = "Não informado"

                if linha["Inscricao_ZSQL"] != "Nao encontrado":
                    lista_registros.append(linha)
    except Exception as e:
        traceback.print_exc()
    return lista_registros


def _obter_pasta_pdfs_sessao(criar_nova=False):
    """Retorna pasta da sessão onde os PDFs originais ficam armazenados."""
    if criar_nova and session.get('pasta_pdfs'):
        pasta_antiga = os.path.join(UPLOAD_FOLDER, session['pasta_pdfs'])
        if os.path.isdir(pasta_antiga):
            shutil.rmtree(pasta_antiga, ignore_errors=True)

    if criar_nova or 'pasta_pdfs' not in session:
        session['pasta_pdfs'] = datetime.now().strftime('%Y%m%d_%H%M%S')

    pasta = os.path.join(UPLOAD_FOLDER, session['pasta_pdfs'])
    os.makedirs(pasta, exist_ok=True)
    return pasta


def _localizar_pdf_na_sessao(pasta_sessao, nome_desejado):
    """Localiza PDF na pasta da sessão (comparação insensível a maiúsculas)."""
    caminho = os.path.join(pasta_sessao, nome_desejado)
    if os.path.isfile(caminho):
        return caminho

    nome_lower = nome_desejado.lower()
    if os.path.isdir(pasta_sessao):
        for arquivo in os.listdir(pasta_sessao):
            if arquivo.lower() == nome_lower:
                return os.path.join(pasta_sessao, arquivo)
    return None


def _coletar_nomes_pdf_features(features):
    """Extrai nomes únicos de PDFs a partir das features filtradas."""
    nomes = set()
    for feat in features:
        for serv in feat.get('properties', {}).get('Servicos', []):
            nome = serv.get('Nome_do_Arquivo')
            if nome:
                nomes.add(os.path.basename(str(nome)))
    return sorted(nomes)


def _copiar_pdfs_para_campo(pasta_sessao, nomes_desejados):
    """Copia PDFs selecionados para pasta pdfs_campo com subpasta datada."""
    pasta_destino = os.path.join(PDFS_CAMPO_FOLDER, datetime.now().strftime('%Y%m%d_%H%M%S'))
    os.makedirs(pasta_destino, exist_ok=True)

    copiados = []
    nao_encontrados = []

    for nome in nomes_desejados:
        origem = _localizar_pdf_na_sessao(pasta_sessao, nome)
        if not origem:
            nao_encontrados.append(nome)
            continue
        destino = os.path.join(pasta_destino, os.path.basename(nome))
        shutil.copy2(origem, destino)
        copiados.append(os.path.basename(nome))

    return pasta_destino, copiados, nao_encontrados

# ==========================================
# ROTAS DE AUTENTICAÇÃO (Usa E_GLOBAL_COMERCIAL)
# ==========================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'usuario' in session:
        return redirect(url_for('index'))

    if request.method == 'POST':
        usuario = request.form.get('usuario', '').strip()
        senha = request.form.get('senha', '').strip()

        try:
            # Conecta no banco E_GLOBAL_COMERCIAL para checar a tabela sec_users
            conn = obter_conexao_sql('E_GLOBAL_COMERCIAL')
            cursor = conn.cursor()
            cursor.execute("SELECT pswd, name FROM dbo.sec_users WHERE login = ? AND active = 'Y'", (usuario,))
            row = cursor.fetchone()
            conn.close()

            if row:
                senha_cadastrada = row[0]
                nome_completo = row[1]

                # Valida senha (suporta texto puro ou hash MD5)
                senha_md5 = hashlib.md5(senha.encode('utf-8')).hexdigest()
                if senha_cadastrada == senha or senha_cadastrada == senha_md5:
                    session['usuario'] = usuario
                    session['nome'] = nome_completo
                    return redirect(url_for('index'))

            return render_template('login.html', erro="Usuário ou senha inválidos, ou conta inativa.")
        except Exception as e:
            print(f"[ERRO LOGIN]: {e}")
            return render_template('login.html', erro="Erro ao conectar no banco de autenticação.")

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ==========================================
# ROTAS DO SISTEMA (Protegidas)
# ==========================================
@app.route('/')
def index():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    return render_template('index.html', nome_usuario=session.get('nome'))

@app.route('/processar_mapa', methods=['POST'])
def processar_mapa():
    if 'usuario' not in session:
        return jsonify({"erro": "Não autorizado. Faça login."}), 401

    if 'arquivos_pdf' not in request.files:
        return jsonify({"erro": "Nenhum arquivo PDF enviado."}), 400
    
    arquivos = request.files.getlist('arquivos_pdf')
    if not arquivos or arquivos[0].filename == '':
        return jsonify({"erro": "Nenhum arquivo selecionado."}), 400

    try:
        todos_dados_planilha = []
        tarefas = []
        pasta_sessao = _obter_pasta_pdfs_sessao(criar_nova=True)

        for arquivo in arquivos:
            if arquivo.filename.lower().endswith('.pdf'):
                nome_original = os.path.basename(arquivo.filename)
                nome_seguro = nome_original.replace('/', '_').replace('\\', '_')
                caminho_temp = os.path.join(pasta_sessao, nome_seguro)
                arquivo.save(caminho_temp)
                tarefas.append((caminho_temp, nome_original))

        with ProcessPoolExecutor() as executor:
            futures = [executor.submit(processar_um_pdf, caminho, nome) for caminho, nome in tarefas]
            for future in as_completed(futures):
                resultado = future.result()
                if resultado:
                    todos_dados_planilha.extend(resultado)

        if not todos_dados_planilha:
            return jsonify({
                "total_planilha": 0,
                "total_sql": 0,
                "total_mapeado": 0,
                "geojson": {"type": "FeatureCollection", "features": []}
            })

        df_pdf = pd.DataFrame(todos_dados_planilha)
        total_planilha = len(df_pdf)

        # ----------------------------------------------------
        # CONEXÃO E CONSULTA NO BANCO DE DADOS SABESP_LIGACOES
        # ----------------------------------------------------
        print("[DEBUG SQL] Conectando ao Banco SABESP_LIGACOES...")
        conn = obter_conexao_sql('SABESP_LIGACOES')
        df_sql = pd.read_sql("SELECT inscricaol, POINT_X, POINT_Y FROM dbo.LIGACOES_CX_SUL", conn)
        conn.close()

        total_sql = len(df_sql)

        df_pdf['chave_pdf'] = df_pdf['Inscricao_ZSQL'].astype(str).str.strip()
        df_sql['chave_sql'] = df_sql['inscricaol'].astype(str).str.strip()

        df_final = pd.merge(df_pdf, df_sql, left_on='chave_pdf', right_on='chave_sql', how='inner')

        features = []
        total_mapeado = 0

        grupos = df_final.groupby(['chave_pdf', 'POINT_X', 'POINT_Y'])

        for (inscricao, lon, lat), grupo in grupos:
            try:
                lat_f = float(lat)
                lon_f = float(lon)
                
                if pd.notna(lat_f) and pd.notna(lon_f) and lat_f != 0 and lon_f != 0:
                    total_mapeado += 1
                    
                    servicos = []
                    nome_consumidor = "Não informado"
                    
                    for _, row in grupo.iterrows():
                        if row.get('Nome_Consumidor') and row.get('Nome_Consumidor') != "Não informado":
                            nome_consumidor = row.get('Nome_Consumidor')
                            
                        servicos.append({
                            "OS_Numero": row.get('OS_Numero'),
                            "Tipo": row.get('Tipo'),
                            "Nome_do_Arquivo": row.get('Nome_do_Arquivo'),
                            "Hidrometro": row.get('Hidrometro')
                        })

                    features.append({
                        "type": "Feature",
                        "geometry": {
                            "type": "Point",
                            "coordinates": [lon_f, lat_f]
                        },
                        "properties": {
                            "Inscricao_ZSQL": str(inscricao),
                            "Nome_Consumidor": nome_consumidor,
                            "Total_Servicos": len(servicos),
                            "Servicos": servicos,
                            "Tipo": servicos[0]["Tipo"] if servicos else "Indefinido"
                        }
                    })
            except Exception:
                continue

        return jsonify({
            "total_planilha": total_planilha,
            "total_sql": total_sql,
            "total_mapeado": total_mapeado,
            "geojson": {"type": "FeatureCollection", "features": features}
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"erro": f"Erro: {str(e)}"}), 500

@app.route('/exportar_excel', methods=['POST'])
def exportar_excel():
    if not requer_login():
        return jsonify({"erro": "Não autorizado."}), 401
    try:
        dados = request.get_json() or {}
        features = dados.get('features', [])
        funcionario_id = dados.get('funcionario_id')

        if not features:
            return jsonify({"erro": "Nenhum registro para exportar."}), 400

        funcionario = None
        if funcionario_id:
            funcionario = obter_funcionario_por_id(funcionario_id)
            if not funcionario:
                return jsonify({"erro": "Funcionário não encontrado."}), 404

        output = gerar_excel_personalizado(features, funcionario, session.get('nome', ''))
        nome_base = 'relatorio_ligacoes'
        if funcionario:
            nome_safe = remover_acentos(funcionario.get('nome', '')).replace(' ', '_')[:30]
            nome_base = f'relatorio_{nome_safe}'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{nome_base}.xlsx'
        )
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/exportar_pdfs_campo', methods=['POST'])
def exportar_pdfs_campo():
    """Copia PDFs da seleção para a pasta pdfs_campo no projeto."""
    if not requer_login():
        return jsonify({"erro": "Não autorizado."}), 401
    try:
        dados = request.get_json(silent=True) or {}
        features = dados.get('features', [])
        if not features:
            return jsonify({"erro": "Nenhuma ligação selecionada para exportar PDFs."}), 400

        if 'pasta_pdfs' not in session:
            return jsonify({
                "erro": "Nenhuma pasta de PDFs carregada nesta sessão. Carregue a pasta e clique em 'Gerar Mapa' novamente."
            }), 400

        pasta_sessao = _obter_pasta_pdfs_sessao()
        if not os.path.isdir(pasta_sessao):
            return jsonify({"erro": "Pasta de PDFs da sessão não encontrada. Reprocesse o mapa."}), 404

        nomes_desejados = _coletar_nomes_pdf_features(features)
        if not nomes_desejados:
            return jsonify({"erro": "Nenhum PDF encontrado nos registros filtrados."}), 400

        pasta_destino, copiados, nao_encontrados = _copiar_pdfs_para_campo(pasta_sessao, nomes_desejados)

        if not copiados:
            return jsonify({
                "erro": "PDFs não encontrados na sessão. Carregue a pasta novamente e reprocesse o mapa.",
                "nao_encontrados": nao_encontrados
            }), 404

        return jsonify({
            "mensagem": f'{len(copiados)} PDF(s) salvos na pasta de campo com sucesso.',
            "pasta": pasta_destino,
            "pasta_relativa": os.path.relpath(pasta_destino, os.path.dirname(os.path.abspath(__file__))),
            "total_copiados": len(copiados),
            "arquivos": copiados,
            "nao_encontrados": nao_encontrados
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"erro": str(e)}), 500


# ==========================================
# MÓDULO DE CADASTRO DE FUNCIONÁRIOS
# ==========================================
@app.route('/funcionarios', methods=['GET'])
def api_listar_funcionarios():
    if not requer_login():
        return jsonify({"erro": "Não autorizado."}), 401
    try:
        incluir_inativos = request.args.get('todos') == '1'
        return jsonify(listar_funcionarios(apenas_ativos=not incluir_inativos))
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/funcionarios', methods=['POST'])
def api_criar_funcionario():
    if not requer_login():
        return jsonify({"erro": "Não autorizado."}), 401
    try:
        funcionario = salvar_funcionario(request.get_json() or {})
        return jsonify({"mensagem": "Funcionário cadastrado com sucesso.", "funcionario": funcionario}), 201
    except ValueError as e:
        return jsonify({"erro": str(e)}), 400
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/funcionarios/<int:funcionario_id>', methods=['PUT'])
def api_atualizar_funcionario(funcionario_id):
    if not requer_login():
        return jsonify({"erro": "Não autorizado."}), 401
    try:
        funcionario = salvar_funcionario(request.get_json() or {}, funcionario_id)
        return jsonify({"mensagem": "Funcionário atualizado com sucesso.", "funcionario": funcionario})
    except ValueError as e:
        return jsonify({"erro": str(e)}), 400
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/funcionarios/<int:funcionario_id>', methods=['DELETE'])
def api_remover_funcionario(funcionario_id):
    if not requer_login():
        return jsonify({"erro": "Não autorizado."}), 401
    try:
        if not remover_funcionario(funcionario_id):
            return jsonify({"erro": "Funcionário não encontrado."}), 404
        return jsonify({"mensagem": "Funcionário removido com sucesso."})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/enviar_excel_funcionario', methods=['POST'])
def enviar_excel_funcionario():
    if not requer_login():
        return jsonify({"erro": "Não autorizado."}), 401
    try:
        dados = request.get_json() or {}
        features = dados.get('features', [])
        funcionario_id = dados.get('funcionario_id')

        if not funcionario_id:
            return jsonify({"erro": "Selecione o funcionário que receberá o relatório."}), 400
        if not features:
            return jsonify({"erro": "Nenhum registro para exportar."}), 400

        funcionario = obter_funcionario_por_id(funcionario_id)
        if not funcionario:
            return jsonify({"erro": "Funcionário não encontrado."}), 404
        if not funcionario.get('email'):
            return jsonify({"erro": "Funcionário sem e-mail cadastrado."}), 400

        output = gerar_excel_personalizado(features, funcionario, session.get('nome', ''))
        nome_safe = remover_acentos(funcionario.get('nome', 'funcionario')).replace(' ', '_')[:30]
        nome_arquivo = f'relatorio_{nome_safe}.xlsx'

        assunto = dados.get('assunto') or f'ENORSUL — Relatório de Ligações para {funcionario.get("nome")}'
        corpo = dados.get('mensagem') or (
            f'Olá {funcionario.get("nome")},\n\n'
            f'Segue em anexo o relatório operacional de ligações gerado pelo sistema ENORSUL.\n'
            f'Total de ligações: {len(features)}\n\n'
            f'Atenciosamente,\n{session.get("nome", "Equipe ENORSUL")}'
        )

        modo_envio = enviar_email_com_anexo(funcionario.get('email'), assunto, corpo, output, nome_arquivo)

        mensagem = f'Relatório enviado com sucesso para {funcionario.get("email")}.'
        if modo_envio == 'outlook':
            mensagem += ' (via Outlook)'

        return jsonify({
            "mensagem": mensagem,
            "funcionario": funcionario.get('nome'),
            "modo_envio": modo_envio
        })
    except ValueError as e:
        return jsonify({"erro": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"erro": f'Falha ao enviar e-mail: {str(e)}'}), 500

def _cor_por_tipo_os(tipo):
    """Retorna cor hexadecimal conforme o tipo de OS."""
    if not tipo:
        return '#6c757d'
    t = str(tipo).upper()
    if 'CORTE' in t:
        return '#dc3545'
    if 'RELIGA' in t:
        return '#28a745'
    if 'LIGA' in t:
        return '#007bff'
    if 'HIDRO' in t or 'TROCA' in t:
        return '#ffc107'
    return '#6c757d'


def _calcular_zoom_mapa(min_x, max_x, min_y, max_y, largura_px, altura_px):
    """Calcula nível de zoom ideal para tiles nítidos no PDF."""
    span_x = max(abs(max_x - min_x), 1.0)
    span_y = max(abs(max_y - min_y), 1.0)
    zoom_x = math.log2((largura_px * 360) / (span_x * 256 / (2 * math.pi)))
    zoom_y = math.log2((altura_px * 360) / (span_y * 256 / (2 * math.pi)))
    zoom = int(min(zoom_x, zoom_y))
    return max(12, min(zoom, 19))


def _ajustar_extent_proporcao(min_x, max_x, min_y, max_y, proporcao_alvo, padding=0.12):
    """Ajusta a extensão do mapa para coincidir com a proporção do quadro do PDF."""
    centro_x = (min_x + max_x) / 2
    centro_y = (min_y + max_y) / 2
    span_x = max(max_x - min_x, 1e-9)
    span_y = max(max_y - min_y, 1e-9)
    proporcao_atual = span_x / span_y

    if proporcao_atual < proporcao_alvo:
        span_x = span_y * proporcao_alvo
    else:
        span_y = span_x / proporcao_alvo

    fator = 1 + padding
    meio_x = span_x * fator / 2
    meio_y = span_y * fator / 2
    return centro_x - meio_x, centro_x + meio_x, centro_y - meio_y, centro_y + meio_y


def _gerar_imagem_mapa(features, largura_px=4800, altura_px=3200):
    """Gera imagem PNG do mapa com a mesma proporção do quadro do PDF."""
    dpi = 300
    proporcao_alvo = largura_px / altura_px
    fig_w = largura_px / dpi
    fig_h = altura_px / dpi

    tipos_contador = Counter()
    for feat in features:
        for serv in feat.get('properties', {}).get('Servicos', []):
            tipos_contador[serv.get('Tipo', 'Desconhecido')] += 1

    paleta_cores = ['#6f42c1', '#28a745', '#007bff', '#dc3545', '#fd7e14', '#e83e8c', '#20c997']
    tipo_para_cor = {
        t_nome: _cor_por_tipo_os(t_nome) if _cor_por_tipo_os(t_nome) != '#6c757d'
        else paleta_cores[idx % len(paleta_cores)]
        for idx, t_nome in enumerate(tipos_contador.keys())
    }

    fig = plt.figure(figsize=(fig_w, fig_h), dpi=dpi, facecolor='#ffffff')
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_facecolor('#eef1f4')

    lons, lats = [], []
    tamanho_ponto = max(35, int(largura_px / 140))

    if HAS_CONTEXTILY:
        transformer = pyproj.Transformer.from_crs('EPSG:4326', 'EPSG:3857', always_xy=True)
        for feat in features:
            lon_raw, lat_raw = feat['geometry']['coordinates']
            x_merc, y_merc = transformer.transform(lon_raw, lat_raw)
            lons.append(x_merc)
            lats.append(y_merc)
            servs = feat.get('properties', {}).get('Servicos', [])
            primeiro_tipo = servs[0].get('Tipo', 'Desconhecido') if servs else 'Desconhecido'
            cor = tipo_para_cor.get(primeiro_tipo, _cor_por_tipo_os(primeiro_tipo))
            ax.scatter(
                x_merc, y_merc, c=cor, s=tamanho_ponto,
                edgecolor='#ffffff', linewidth=1.4, zorder=5, alpha=0.98
            )
            ax.scatter(
                x_merc, y_merc, c='none', s=tamanho_ponto,
                edgecolor='#1a1a1a', linewidth=0.9, zorder=6, alpha=1.0
            )
    else:
        for feat in features:
            lon_raw, lat_raw = feat['geometry']['coordinates']
            lons.append(lon_raw)
            lats.append(lat_raw)
            servs = feat.get('properties', {}).get('Servicos', [])
            primeiro_tipo = servs[0].get('Tipo', 'Desconhecido') if servs else 'Desconhecido'
            cor = tipo_para_cor.get(primeiro_tipo, _cor_por_tipo_os(primeiro_tipo))
            ax.scatter(
                lon_raw, lat_raw, c=cor, s=tamanho_ponto,
                edgecolor='#ffffff', linewidth=1.4, zorder=5, alpha=0.98
            )
            ax.scatter(
                lon_raw, lat_raw, c='none', s=tamanho_ponto,
                edgecolor='#1a1a1a', linewidth=0.9, zorder=6, alpha=1.0
            )

    span_minimo = 2500.0 if HAS_CONTEXTILY else 0.012
    if lons and lats:
        min_x, max_x = min(lons), max(lons)
        min_y, max_y = min(lats), max(lats)
        if max_x - min_x < span_minimo:
            min_x -= span_minimo / 2
            max_x += span_minimo / 2
        if max_y - min_y < span_minimo:
            min_y -= span_minimo / 2
            max_y += span_minimo / 2
        x0, x1, y0, y1 = _ajustar_extent_proporcao(min_x, max_x, min_y, max_y, proporcao_alvo)
        ax.set_xlim(x0, x1)
        ax.set_ylim(y0, y1)

    mapa_carregado = False
    if HAS_CONTEXTILY and lons and lats:
        try:
            xlim = ax.get_xlim()
            ylim = ax.get_ylim()
            zoom = _calcular_zoom_mapa(xlim[0], xlim[1], ylim[0], ylim[1], largura_px, altura_px)
            ctx.add_basemap(
                ax,
                crs='EPSG:3857',
                source=ctx.providers.OpenStreetMap.Mapnik,
                zoom=zoom,
                attribution=False,
                reset_extent=False
            )
            mapa_carregado = True
        except Exception as exc:
            print(f'[AVISO MAPA] Basemap indisponível: {exc}')

    if not mapa_carregado:
        ax.grid(True, linestyle='-', linewidth=0.5, color='#d0d5db', alpha=0.85, zorder=1)

    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_aspect('equal', adjustable='box')
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_color('#94a3b8')
        spine.set_linewidth(0.8)

    img_buffer = io.BytesIO()
    fig.savefig(
        img_buffer,
        format='png',
        dpi=dpi,
        pad_inches=0,
        facecolor='#ffffff',
        transparent=False
    )
    plt.close(fig)
    img_buffer.seek(0)
    return img_buffer, tipos_contador, tipo_para_cor


def _truncar_texto(texto, fonte, tamanho, largura_max):
    """Encurta texto para caber na largura disponível."""
    texto = str(texto or '—')
    if stringWidth(texto, fonte, tamanho) <= largura_max:
        return texto
    while len(texto) > 3 and stringWidth(texto + '…', fonte, tamanho) > largura_max:
        texto = texto[:-1]
    return texto + '…'


def _desenhar_norte_escala(c, x, y, escala_texto='1:7.000'):
    """Desenha rosa dos ventos e indicador de escala no canto inferior esquerdo."""
    box_w = 24 * mm
    box_h = 30 * mm
    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor('#475569'))
    c.setLineWidth(0.7)
    c.roundRect(x, y, box_w, box_h, 2, fill=1, stroke=1)

    cx = x + box_w / 2
    base_y = y + 10 * mm
    c.setFillColor(colors.HexColor('#0f172a'))
    c.setLineWidth(1.1)
    c.line(cx, base_y, cx, base_y + 11 * mm)
    c.line(cx - 3.5 * mm, base_y + 7 * mm, cx, base_y + 12.5 * mm)
    c.line(cx + 3.5 * mm, base_y + 7 * mm, cx, base_y + 12.5 * mm)
    c.setFont('Helvetica-Bold', 8)
    c.drawCentredString(cx, base_y + 14 * mm, 'N')

    bar_y = y + 4 * mm
    c.setFillColor(colors.HexColor('#0f172a'))
    c.rect(x + 4 * mm, bar_y, 6 * mm, 1.4 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.rect(x + 10 * mm, bar_y, 6 * mm, 1.4 * mm, fill=1, stroke=0)
    c.setFillColor(colors.HexColor('#0f172a'))
    c.rect(x + 16 * mm, bar_y, 6 * mm, 1.4 * mm, fill=1, stroke=0)

    c.setFont('Helvetica', 5.5)
    c.drawCentredString(cx, y + 1.5 * mm, escala_texto)


def _desenhar_legenda_mapa(c, x, y, largura, tipos_contador, tipo_para_cor, total_ligacoes):
    """Legenda flutuante no canto superior direito do mapa."""
    total_servicos = sum(tipos_contador.values())
    itens = sorted(tipos_contador.items(), key=lambda item: item[1], reverse=True)
    max_itens = min(len(itens), 7)
    altura = 10 * mm + max_itens * 6.5 * mm + 12 * mm

    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor('#475569'))
    c.setLineWidth(0.7)
    c.roundRect(x, y, largura, altura, 2, fill=1, stroke=1)

    c.setFillColor(colors.HexColor('#c82333'))
    c.rect(x, y + altura - 7 * mm, largura, 7 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 7)
    c.drawString(x + 3 * mm, y + altura - 5 * mm, 'LEGENDA TSS')

    cursor_y = y + altura - 11 * mm
    c.setFont('Helvetica', 6.5)
    for tipo, qtd in itens[:max_itens]:
        cor = tipo_para_cor.get(tipo, _cor_por_tipo_os(tipo))
        c.setFillColor(colors.HexColor(cor))
        c.setStrokeColor(colors.HexColor('#1a1a1a'))
        c.setLineWidth(0.4)
        c.circle(x + 4.5 * mm, cursor_y, 1.8 * mm, fill=1, stroke=1)
        c.setFillColor(colors.HexColor('#0f172a'))
        nome = _truncar_texto(tipo, 'Helvetica', 6.5, largura - 18 * mm)
        c.drawString(x + 8 * mm, cursor_y - 1.2 * mm, nome)
        c.setFillColor(colors.HexColor('#64748b'))
        c.drawRightString(x + largura - 3 * mm, cursor_y - 1.2 * mm, str(qtd))
        cursor_y -= 6.5 * mm

    c.setStrokeColor(colors.HexColor('#e2e8f0'))
    c.setLineWidth(0.3)
    c.line(x + 2 * mm, y + 7 * mm, x + largura - 2 * mm, y + 7 * mm)
    c.setFillColor(colors.HexColor('#64748b'))
    c.setFont('Helvetica', 5.5)
    c.drawString(x + 3 * mm, y + 2.5 * mm, f'{total_ligacoes} ligações • {total_servicos} OS')


def _desenhar_carimbo(c, x, y, largura, altura, meta, session, total_ligacoes):
    """Carimbo técnico inferior com metadados em grade 3x3."""
    col_w = largura / 3
    row_h = altura / 3

    c.setStrokeColor(colors.HexColor('#334155'))
    c.setLineWidth(0.6)
    c.rect(x, y, largura, altura, fill=0, stroke=1)
    for i in range(1, 3):
        c.line(x + col_w * i, y, x + col_w * i, y + altura)
        c.line(x, y + row_h * i, x + largura, y + row_h * i)

    campos = [
        ('DIRETORIA', meta.get('diretoria', 'DIRETORIA METROPOLITANA')),
        ('UNIDADE DE NEGÓCIO', meta.get('unidade', 'UNIDADE DE NEGÓCIO CENTRO')),
        ('REFERÊNCIA', meta.get('referencia', '—')),
        ('TSS', meta.get('tss', '—')),
        ('ESCALA', meta.get('escala', '1:7.000')),
        ('DATA', meta.get('data', '—')),
        ('ELABORADO', meta.get('elaborado') or session.get('nome', '—')),
        ('STATUS DO DESENHO', meta.get('status_desenho', 'Preliminar')),
        ('TOTAL NA ÁREA', f'{total_ligacoes} ligações'),
    ]

    for idx, (rotulo, valor) in enumerate(campos):
        col = idx % 3
        row = 2 - (idx // 3)
        cx = x + col * col_w + 3 * mm
        cy = y + row * row_h + row_h - 5 * mm
        c.setFillColor(colors.HexColor('#64748b'))
        c.setFont('Helvetica-Bold', 5.5)
        c.drawString(cx, cy, rotulo)
        c.setFillColor(colors.HexColor('#0f172a'))
        c.setFont('Helvetica', 7)
        c.drawString(cx, cy - 4 * mm, _truncar_texto(valor, 'Helvetica', 7, col_w - 6 * mm))


def _montar_layout_pdf(formato):
    """Calcula dimensões fixas do layout cartográfico."""
    pagesize = landscape(A3 if formato == 'A3' else A4)
    width, height = pagesize
    margem = 8 * mm
    cabecalho_h = 20 * mm
    carimbo_h = 30 * mm
    rodape_h = 5 * mm
    area_util_w = width - 2 * margem
    area_util_h = height - 2 * margem
    map_h = area_util_h - cabecalho_h - carimbo_h - rodape_h - 4 * mm
    map_w = area_util_w
    return {
        'pagesize': pagesize,
        'width': width,
        'height': height,
        'margem': margem,
        'cabecalho_h': cabecalho_h,
        'carimbo_h': carimbo_h,
        'rodape_h': rodape_h,
        'map_w': map_w,
        'map_h': map_h,
        'map_x': margem,
        'map_y': margem + carimbo_h + rodape_h,
        'cabecalho_y': height - margem - cabecalho_h,
        'carimbo_y': margem,
    }


def _desenhar_cabecalho_pdf(c, layout, meta, titulo):
    """Cabeçalho institucional limpo."""
    x = layout['margem']
    y = layout['cabecalho_y']
    w = layout['map_w']
    h = layout['cabecalho_h']

    c.setFillColor(colors.HexColor('#ffffff'))
    c.setStrokeColor(colors.HexColor('#cbd5e1'))
    c.setLineWidth(0.6)
    c.rect(x, y, w, h, fill=1, stroke=1)

    c.setFillColor(colors.HexColor('#c82333'))
    c.rect(x, y, 4 * mm, h, fill=1, stroke=0)

    c.setFillColor(colors.HexColor('#c82333'))
    c.setFont('Helvetica-Bold', 11)
    c.drawString(x + 7 * mm, y + h - 7 * mm, 'ENORSUL')

    c.setFillColor(colors.HexColor('#0f172a'))
    c.setFont('Helvetica-Bold', 10)
    c.drawString(x + 7 * mm, y + 3.5 * mm, _truncar_texto(titulo, 'Helvetica-Bold', 10, w - 80 * mm))

    c.setFillColor(colors.HexColor('#64748b'))
    c.setFont('Helvetica', 7)
    c.drawRightString(x + w - 4 * mm, y + h - 6 * mm, f"Folha: {meta.get('folha', '01/01')}")
    c.drawRightString(x + w - 4 * mm, y + 4 * mm, 'Monitoramento Operacional de Ligações')


def _renderizar_pdf_mapa(c, layout, img_buffer, tipos_contador, tipo_para_cor, meta, session, total_ligacoes):
    """Renderiza quadro do mapa com imagem encaixada sem distorção."""
    map_x = layout['map_x']
    map_y = layout['map_y']
    map_w = layout['map_w']
    map_h = layout['map_h']
    borda = 2 * mm
    img_x = map_x + borda
    img_y = map_y + borda
    img_w = map_w - 2 * borda
    img_h = map_h - 2 * borda

    c.setFillColor(colors.HexColor('#e2e8f0'))
    c.setStrokeColor(colors.HexColor('#334155'))
    c.setLineWidth(1.2)
    c.rect(map_x, map_y, map_w, map_h, fill=1, stroke=1)

    c.setFillColor(colors.HexColor('#cbd5e1'))
    c.rect(img_x, img_y, img_w, img_h, fill=1, stroke=0)

    c.drawImage(
        ImageReader(img_buffer),
        img_x,
        img_y,
        img_w,
        img_h,
        preserveAspectRatio=False,
        anchor='sw',
        mask='auto'
    )

    c.setStrokeColor(colors.HexColor('#475569'))
    c.setLineWidth(0.6)
    c.rect(img_x, img_y, img_w, img_h, fill=0, stroke=1)

    if tipos_contador:
        _desenhar_legenda_mapa(
            c,
            img_x + img_w - 66 * mm,
            img_y + img_h - 52 * mm,
            62 * mm,
            tipos_contador,
            tipo_para_cor,
            total_ligacoes
        )

    _desenhar_norte_escala(c, img_x + 4 * mm, img_y + 4 * mm, meta.get('escala', '1:7.000'))

    _desenhar_carimbo(
        c,
        layout['margem'],
        layout['carimbo_y'],
        layout['map_w'],
        layout['carimbo_h'],
        meta,
        session,
        total_ligacoes
    )

    c.setFillColor(colors.HexColor('#94a3b8'))
    c.setFont('Helvetica', 5.5)
    c.drawCentredString(
        layout['width'] / 2,
        layout['margem'] + 1 * mm,
        'Documento gerado automaticamente pelo sistema ENORSUL — Mapa 300 DPI'
    )


@app.route('/exportar_pdf', methods=['POST'])
def exportar_pdf():
    if 'usuario' not in session:
        return jsonify({"erro": "Não autorizado."}), 401
    try:
        dados = request.get_json() or {}
        features = dados.get('features', [])
        meta = dados.get('metadata', {})

        if not features:
            return jsonify({"erro": "Nenhum ponto disponível para gerar o mapa no PDF."}), 400

        formato = meta.get('formato', 'A3').upper()
        layout = _montar_layout_pdf(formato)
        width, height = layout['width'], layout['height']

        largura_px = max(4000, int(layout['map_w'] * 350 / 72))
        altura_px = max(2800, int(layout['map_h'] * 350 / 72))

        img_buffer, tipos_contador, tipo_para_cor = _gerar_imagem_mapa(
            features,
            largura_px=largura_px,
            altura_px=altura_px
        )

        titulo = meta.get('titulo', 'Mapa Comercial — Região Selecionada')
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=layout['pagesize'])

        c.setFillColor(colors.white)
        c.rect(0, 0, width, height, fill=1, stroke=0)

        c.setStrokeColor(colors.HexColor('#1e293b'))
        c.setLineWidth(1.5)
        c.rect(layout['margem'] - 2, layout['margem'] - 2,
               width - 2 * layout['margem'] + 4, height - 2 * layout['margem'] + 4)

        _desenhar_cabecalho_pdf(c, layout, meta, titulo)
        _renderizar_pdf_mapa(
            c, layout, img_buffer, tipos_contador, tipo_para_cor,
            meta, session, len(features)
        )

        c.save()
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name='mapa_comercial_relatorio.pdf'
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({"erro": str(e)}), 500

if __name__ == '__main__':
    inicializar_tabela_funcionarios()
    app.run(debug=True, port=5000)